from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

def format_excel(file_path):
    # Verificar se o arquivo existe
    if not os.path.exists(file_path):
        print(f"Arquivo '{file_path}' não encontrado")
        return

    wb = load_workbook(file_path)
    ws = wb.active

    # Determinar a última linha com dados (sem incluir somatório, se presente)
    last_data_row = ws.max_row

    # Definir os grupos de colunas por letra
    grupo_geral = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
        "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA"
    ]
    grupo_duty = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
    grupo_freight = ["T", "U", "V", "W", "X", "Y", "Z", "AA"]

    # Cores
    verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    laranja = PatternFill(start_color='F5CA48', end_color='F5CA48', fill_type='solid')
    azul_claro = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    sem_fundo = PatternFill(fill_type=None)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Regra 1/1: BALANCE = 0, BANK IN DUTY ≠ "?", ≠ "PAID IN CHINA", BALANCE FREIGHT = 0, CONFIRMATION = "CONFIRMADO" → fundo verde no grupo geral
    for col in grupo_geral:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($M2=0, $N2<>"?", $N2<>"PAID IN CHINA", $W2=0, $O2="CONFIRMADO")'], fill=verde)
        )

    # Regra 1/2: BALANCE = 0, BANK IN DUTY = "PAID IN CHINA" → fundo laranja no grupo geral
    for col in grupo_geral:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($M2=0, $N2="PAID IN CHINA")'], fill=laranja)
        )

    # Regra 2: BALANCE = 0, BANK IN DUTY ≠ "?", ≠ "PAID IN CHINA", BALANCE FREIGHT = 0, CONFIRMATION = "CONFIRMADO" → fundo verde no grupo duty
    for col in grupo_duty:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($M2=0, $N2<>"?", $N2<>"PAID IN CHINA", $W2=0, $O2="CONFIRMADO")'], fill=verde)
        )

    # Regra 3: BALANCE = 0, BANK IN DUTY = "PAID IN CHINA" → fundo laranja no grupo duty
    for col in grupo_duty:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($M2=0, $N2="PAID IN CHINA")'], fill=laranja)
        )

    # Regra 4: BALANCE FREIGHT = 0 e NOTA FREIGHT ≠ "PAID TO JUPITER" → fundo verde no grupo freight
    for col in grupo_freight:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($W2=0, $Y2<>"PAID TO JUPITER")'], fill=verde)
        )

    # Regra 5: BALANCE FREIGHT = 0 e NOTA FREIGHT = "PAID TO JUPITER" → fundo azul claro no grupo freight
    for col in grupo_freight:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['AND($W2=0, $Y2="PAID TO JUPITER")'], fill=azul_claro)
        )

    # Regra 6: BALANCE FREIGHT > 0 → fundo sem cor no grupo freight
    for col in grupo_freight:
        ws.conditional_formatting.add(
            f"{col}2:{col}{last_data_row}",
            FormulaRule(formula=['$W2>0'], fill=sem_fundo)
        )

    # Aplicar fórmulas no Excel
    column_names = [cell.value for cell in ws[1]]
    
    # Converter coluna PHONE NUMBER para texto (evitar formatação como número decimal)
    if "PHONE NUMBER" in column_names:
        phone_col_idx = column_names.index("PHONE NUMBER") + 1
        for row in range(2, last_data_row + 1):
            cell = ws.cell(row, phone_col_idx)
            if cell.value is not None:
                # Converter para texto e remover .0 se for float
                phone_str = str(cell.value)
                if phone_str.endswith('.0'):
                    phone_str = phone_str[:-2]  # Remover '.0' do final
                cell.value = phone_str
                cell.number_format = '@'  # Formato de texto no Excel
    
    for row in range(2, last_data_row + 1):
        cbm_idx = column_names.index("CBM") + 1
        duty_prepaid_idx = column_names.index("DUTY PREPAID") + 1
        bank_in_duty_idx = column_names.index("BANK IN DUTY") + 1
        confirmation_idx = column_names.index("CONFIRMATION") + 1
        unit_cbm_duty_idx = column_names.index("UNIT CBM DUTY") + 1
        unit_cbm_freight_idx = column_names.index("UNIT CBM FREIGHT") + 1
        amount_duty_idx = column_names.index("AMOUNT DUTY") + 1
        paid_idx = column_names.index("PAID") + 1
        balance_idx = column_names.index("BALANCE") + 1
        amount_freight_idx = column_names.index("AMOUNT FREIGHT") + 1
        paid_freight_idx = column_names.index("PAID FREIGHT") + 1
        balance_freight_idx = column_names.index("BALANCE FREIGHT") + 1
        status_idx = column_names.index("STATUS") + 1

        cbm = float(ws.cell(row, cbm_idx).value or 0)
        duty = float(ws.cell(row, duty_prepaid_idx).value or 0)

        if cbm == 0:
            ws.cell(row, bank_in_duty_idx).value = "REPOSIÇÃO"
            ws.cell(row, confirmation_idx).value = "CONFIRMADO"
        elif duty > 0:
            ws.cell(row, bank_in_duty_idx).value = "PAID IN CHINA"
            ws.cell(row, confirmation_idx).value = "CONFIRMADO"

        ws.cell(row, amount_duty_idx).value = f"={chr(64 + cbm_idx)}{row}*{chr(64 + unit_cbm_duty_idx)}{row}"
        ws.cell(row, amount_freight_idx).value = f"={chr(64 + cbm_idx)}{row}*{chr(64 + unit_cbm_freight_idx)}{row}"
        ws.cell(row, balance_idx).value = f"={chr(64 + amount_duty_idx)}{row}-{chr(64 + duty_prepaid_idx)}{row}-{chr(64 + paid_idx)}{row}"
        ws.cell(row, balance_freight_idx).value = f"={chr(64 + amount_freight_idx)}{row}-{chr(64 + paid_freight_idx)}{row}"
        # STATUS será definido posteriormente pela lógica de merge por cliente

    # Adicionar listas dropdown para BANK IN DUTY (coluna N) e CONFIRMATION (coluna O)
    bank_options = "?,BCI BOSS,BIM BOSS,BCI JUPITER,BIM JUPITER,STB JUPITER,NED JUPITER,EMOLA BOSS,PAID IN CHINA,REPOSIÇÃO,COTACAO"
    confirmation_options = "?,PARCIAL,CONFIRMADO,RE-VERIFICANDO,SEM COMPROVATIVO,COMPROVATIVO ERRADO"
    
    dv_bank = DataValidation(type="list", formula1=f'"{bank_options}"', allow_blank=True)
    dv_bank.add(f"N2:N{last_data_row}")
    ws.add_data_validation(dv_bank)

    dv_confirmation = DataValidation(type="list", formula1=f'"{confirmation_options}"', allow_blank=True)
    dv_confirmation.add(f"O2:O{last_data_row}")
    ws.add_data_validation(dv_confirmation)

    # Adicionar lista dropdown para BANK IN FREIGHT (coluna U), para consistência
    dv_bank_freight = DataValidation(type="list", formula1=f'"{bank_options}"', allow_blank=True)
    dv_bank_freight.add(f"X2:X{last_data_row}")
    ws.add_data_validation(dv_bank_freight)

    # Adicionar linha de total
    total_row = last_data_row + 1
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 1).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    
    numeric_columns = ["CBM", "UNIT CBM DUTY", "DUTY PREPAID", "AMOUNT DUTY", "PAID", "BALANCE", 
                       "UNIT CBM FREIGHT", "AMOUNT FREIGHT", "PAID FREIGHT", "BALANCE FREIGHT", "PACKAGES"]
    for col in numeric_columns:
        idx = column_names.index(col) + 1
        ws.cell(total_row, idx).value = f"=SUM({chr(64 + idx)}2:{chr(64 + idx)}{last_data_row})"
        ws.cell(total_row, idx).font = Font(bold=True)
        ws.cell(total_row, idx).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    # Formatação adicional: bordas
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f"A1:{get_column_letter(len(column_names))}{total_row}"]:
        for cell in row:
            cell.border = thin_border

    # Formatação adicional: alinhamento
    for row in ws[f"A1:{get_column_letter(len(column_names))}{total_row}"]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatação adicional: ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[column].width = adjusted_width

    # Formatação adicional: cabeçalhos
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Formatação condicional adicional: valores negativos em vermelho
    for col in ["BALANCE", "BALANCE FREIGHT"]:
        col_idx = chr(64 + column_names.index(col))
        ws.conditional_formatting.add(
            f"{col_idx}2:{col_idx}{last_data_row}",
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
        )

    # --- MERGE DE CÉLULAS PARA CLIENTES COM MÚLTIPLAS ORDENS ---
    # Primeiro, avaliar todas as fórmulas de STATUS antes de fazer merge
    status_col_idx = column_names.index("STATUS") + 1
    
    # Forçar o cálculo/avaliação de todas as células de STATUS
    # (as fórmulas já foram inseridas anteriormente)
    
    # Identificar grupos de clientes baseado em ID CODE
    id_code_col_idx = column_names.index("ID CODE") + 1
    no_col_idx = column_names.index("NO") + 1
    
    client_groups = []
    current_group_start = 2
    current_id_code = ws.cell(2, id_code_col_idx).value
    current_no = ws.cell(2, no_col_idx).value
    
    # Percorrer todas as linhas para identificar grupos
    # Usar NO como identificador primário (porque pode estar vazio em linhas merged)
    for row in range(3, last_data_row + 1):
        row_no = ws.cell(row, no_col_idx).value
        
        # Se NO estiver vazio, é a mesma pessoa (linhas com múltiplas ordens)
        if row_no is None or row_no == "":
            continue
        
        # Se mudou de cliente (NO diferente)
        if row_no != current_no:
            # Fechou um grupo, adicionar à lista
            client_groups.append({
                'start_row': current_group_start,
                'end_row': row - 1,
                'no': current_no
            })
            current_group_start = row
            current_no = row_no
    
    # Adicionar o último grupo
    client_groups.append({
        'start_row': current_group_start,
        'end_row': last_data_row,
        'no': current_no
    })
    
    # Processar cada grupo de cliente
    balance_col_idx = column_names.index("BALANCE") + 1
    balance_freight_col_idx = column_names.index("BALANCE FREIGHT") + 1
    confirmation_col_idx = column_names.index("CONFIRMATION") + 1
    bank_in_duty_col_idx = column_names.index("BANK IN DUTY") + 1
    
    for group in client_groups:
        start_row = group['start_row']
        end_row = group['end_row']
        
        # Construir fórmula Excel que verifica TODAS as linhas do grupo
        # Fórmula: =IF(AND(COUNTIF(M{start}:M{end},0)={count}, COUNTIF(T{start}:T{end},0)={count}, COUNTIF(O{start}:O{end},"CONFIRMADO")={count}, COUNTIF(N{start}:N{end},"?")=0, COUNTIF(N{start}:N{end},"")=0), "PODE LEVANTAR", "")
        
        balance_col = chr(64 + balance_col_idx)
        balance_freight_col = chr(64 + balance_freight_col_idx)
        confirmation_col = chr(64 + confirmation_col_idx)
        bank_in_duty_col = chr(64 + bank_in_duty_col_idx)
        
        num_orders = end_row - start_row + 1
        
        # Fórmula que verifica:
        # 1. Todas as células BALANCE = 0
        # 2. Todas as células BALANCE FREIGHT = 0
        # 3. Todas as células CONFIRMATION = "CONFIRMADO"
        # 4. Nenhuma célula BANK IN DUTY = "?"
        # 5. Nenhuma célula BANK IN DUTY vazia
        status_formula = (
            f'=IF(AND('
            f'COUNTIF({balance_col}{start_row}:{balance_col}{end_row},0)={num_orders},'
            f'COUNTIF({balance_freight_col}{start_row}:{balance_freight_col}{end_row},0)={num_orders},'
            f'COUNTIF({confirmation_col}{start_row}:{confirmation_col}{end_row},"CONFIRMADO")={num_orders},'
            f'COUNTIF({bank_in_duty_col}{start_row}:{bank_in_duty_col}{end_row},"?")=0,'
            f'COUNTIF({bank_in_duty_col}{start_row}:{bank_in_duty_col}{end_row},"")=0'
            f'),"PODE LEVANTAR","")'
        )
        
        # Definir formatação condicional para STATUS
        status_fill_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        status_font_green = Font(bold=True, color='006400')
        
        # Fazer merge das colunas: STATUS, NO, ID CODE, NAME, PHONE NUMBER
        merge_columns = ["STATUS", "NO", "ID CODE", "NAME", "PHONE NUMBER"]
        
        for col_name in merge_columns:
            if col_name in column_names:
                col_idx = column_names.index(col_name) + 1
                
                # Fazer merge se houver múltiplas linhas
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                
                # Definir o valor apropriado
                if col_name == "STATUS":
                    # Inserir fórmula Excel dinâmica
                    ws.cell(start_row, col_idx).value = status_formula
                
                # Centralizar verticalmente
                ws.cell(start_row, col_idx).alignment = Alignment(horizontal='center', vertical='center')
    
    # Aplicar formatação condicional para STATUS = "PODE LEVANTAR"
    # Usar formatação condicional do Excel para cor verde
    status_col_letter = chr(64 + column_names.index("STATUS") + 1)
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    green_font = Font(bold=True, color='006400')
    
    ws.conditional_formatting.add(
        f'{status_col_letter}2:{status_col_letter}{last_data_row}',
        CellIsRule(operator='equal', formula=['"PODE LEVANTAR"'], fill=green_fill, font=green_font)
    )
    
    # --- PROTEÇÃO DA PLANILHA ---
    # Colunas protegidas: NO, ID CODE, NAME, PHONE NUMBER, ORDER NUMBER, CARGO DESCRIPTION (PACKAGES), CBM, UNIT CBM DUTY
    protected_columns = ["NO", "ID CODE", "NAME", "PHONE NUMBER", "ORDER NUMBER", 
                         "CARGO DESCRIPTION (PACKAGES)", "CBM", "UNIT CBM DUTY"]
    
    # Desbloquear todas as células primeiro
    for row in ws.iter_rows(min_row=2, max_row=last_data_row):
        for cell in row:
            cell.protection = cell.protection.copy(locked=False)
    
    # Bloquear apenas as colunas especificadas
    for col_name in protected_columns:
        if col_name in column_names:
            col_idx = column_names.index(col_name) + 1
            for row in range(2, last_data_row + 1):
                ws.cell(row, col_idx).protection = ws.cell(row, col_idx).protection.copy(locked=True)
    
    # Proteger a planilha com senha 792721
    ws.protection.sheet = True
    ws.protection.password = '792721'
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False

    # Salvar a versão final com todas as formatações aplicadas
    wb.save(file_path)
